import pyodbc
import pandas as pd
import re
import os
from openpyxl import load_workbook

outdir = os.path.dirname(os.path.realpath(__file__))
issql = re.compile('.*\.sql$')   #Matches ending in .sql

cnxn = pyodbc.connect('Driver={SQL Server};' # server type driver 
                      'Server=;' #server 
                      'DATABASE=;' #database
                        'Trusted_Connection=yes;'
                        'UID=;' #user id
                        'PWD=') #password


with pd.ExcelWriter(str(outdir)+'\export.xlsx', engine='openpyxl', mode= 'a') as writer:
    writer.book = load_workbook('export.xlsx') ## this will error if does not exist need error handling for this
    for subdir, dir, files in os.walk(outdir): # search dirctory for sql scripts
        for file in files:
            if issql.match(file):
                readfile = open(file, 'r')
                df = pd.read_sql_query(readfile.read(), cnxn) ## get results
                sheet_name = str(file).replace('.sql','') ## parse the .sql for tab name
                if sheet_name in writer.book.sheetnames: ## remove if exist otherwise just write  
                    idx = writer.book.sheetnames.index(sheet_name)
                    writer.book.remove(writer.book.worksheets[idx])
                df.to_excel(writer, sheet_name=sheet_name, index=None, header=True, startrow=3)